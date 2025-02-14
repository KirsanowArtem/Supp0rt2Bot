import asyncio
import re

import aiohttp
import config
import nest_asyncio
import os
import pytz
import threading
import json
import pandas as pd
import requests

import telegram.error
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, ChatPermissions, \
    BotCommand, BotCommandScopeDefault, BotCommandScopeChat, Bot
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, CallbackContext, ContextTypes
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
import hashlib

from aiocron import crontab
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill

from apscheduler.schedulers.background import BackgroundScheduler


from telegram.ext import Application

from gevent import monkey

nest_asyncio.apply()

global muted_users


scheduler = BackgroundScheduler(timezone="Europe/Kiev")

EXCEL_FILE = "user_data_export.xlsx"

application = None

app = Flask(__name__)
from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import hashlib
import json

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Ключ для сессий

DATA_FILE = "data.json"
CHATS_FILE = "chats.json"

<<<<<<< HEAD
SITE_URL = "http://127.0.0.1:5000"
=======
>>>>>>> 116c7b8 (Первый коммит)


# Хешируем пароль "12" через SHA256
VALID_USERNAME = "Skeleton"
VALID_PASSWORD_HASH = hashlib.sha256("12".encode()).hexdigest()



def load_users(file_path=DATA_FILE):
    """Загружает пользователей и формирует список с доп. инфо."""
    data = load_data(file_path)
    users = []
    for user in data["users"]:
        username = user.get("username", "")
        avatar_url = f"https://t.me/i/userpic/320/{username}.jpg" if username else "https://via.placeholder.com/50"

        mute_status = user.get("mute", False)
        mute_end_date = user.get("mute_end", "None")

        # Определяем статус мута
        if mute_status and mute_end_date != "None":
            status = f"🔴 В муте (до {mute_end_date})"
        else:
            status = "🟢 Размучен"

        users.append({
            "id": user["id"],
            "second_name": user["second_name"],  # Используем second_name
            "username": username,
            "avatar": avatar_url,
            "status": status,
            "rating": user.get("rating", 0),
            "mute_end": mute_end_date
        })
    return users


def get_statistics():
    """Возвращает количество пользователей и среднюю оценку."""
    data = load_data(DATA_FILE)
    total_users = len(data["users"])
    total_score = data.get("total_score", 0)
    num_ratings = data.get("num_of_ratings", 1)
    avg_rating = round(total_score / num_ratings, 1) if num_ratings > 0 else 0
    return total_users, avg_rating


def load_data(filename):
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)

def load_chats():
    try:
        with open(CHATS_FILE, "r", encoding="utf-8") as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

# Функция для сохранения чатов в файл
def save_chats(chats):
    with open(CHATS_FILE, "w", encoding="utf-8") as file:
        json.dump(chats, file, ensure_ascii=False, indent=4)

# Сохранение сообщений в файл
def save_message_to_chat(message_id, user_id, text):
    chats = load_chats()
    if message_id not in chats:
        chats[message_id] = {
            "user_id": user_id,
            "messages": []
        }
    chats[message_id]["messages"].append({
        "message_type": "text",
        "text": text
    })
    save_chats(chats)

@app.route('/get_chat_messages')
def get_chat_messages():
    user_id = request.args.get('userId')

    # Загружаем данные из файла
    with open(CHATS_FILE, 'r', encoding='utf-8') as file:
        chats_data = json.load(file)

    # Получаем сообщения для конкретного user_id
    if str(user_id) in chats_data:
        messages = chats_data[str(user_id)]['messages']
    else:
        messages = []

    # Формируем ответ
    formatted_messages = [
        {"username": message["username"], "message": message["message"], "time_sent": message["time_sent"]}
        for message in messages
    ]

    # Отправляем их в формате JSON
    return jsonify({"messages": formatted_messages})




@app.route('/')
def index():
    data = load_data(DATA_FILE)
    users = data['users']
    total_users = len(users)
    avg_rating = sum(user['rating'] for user in users) / total_users if total_users > 0 else 0
    return render_template("main.html", users=users, total_users=total_users, avg_rating=avg_rating)


@app.route('/update_name', methods=['POST'])
def update_name():
    data = request.get_json()
    user_id = data['userId']
    new_name = data['newName']

    # Загружаем данные пользователей
    users_data = load_data(DATA_FILE)
    users = users_data['users']

    # Ищем пользователя по ID и меняем имя
    user_found = False
    for user in users:
        if str(user['id']) == str(user_id):
            user['second_name'] = new_name
            user_found = True
            break

    if user_found:
        save_data(users_data)  # Сохраняем данные
        return jsonify({'success': True, 'new_name': new_name})
    else:
        return jsonify({'success': False})


@app.route("/login", methods=["GET", "POST"])
def login():
    """Страница входа."""
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        hashed_password = hashlib.sha256(password.encode()).hexdigest()

        if username == VALID_USERNAME and hashed_password == VALID_PASSWORD_HASH:
            session["logged_in"] = True
            return redirect(url_for("index"))

    return render_template("login.html")


@app.route("/logout", methods=["POST"])
def logout():
    """Выход из аккаунта."""
    session.pop("logged_in", None)
    return redirect(url_for("login"))


@app.route('/update_chat', methods=['POST'])
def update_chat():
    data = request.json  # Получаем JSON
    if not data:
        return jsonify({"error": "No data received"}), 400

    print("📩 Новое сообщение:", data)

    # ТУТ добавь код обновления страницы на твоем сайте (например, с WebSocket)

    return jsonify({"status": "ok"}), 200


def send_message(chat_id, text):
    url = f"https://api.telegram.org/bot{BOTTOCEN}/sendMessage"
    data = {"chat_id": chat_id, "text": text}
    response = requests.post(url, json=data)
    return response.json()

<<<<<<< HEAD
=======
def get_user_id_by_username(username):
    with open(CHATS_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    for user in data.get("users", []):
        if user.get("username") == username:
            return user.get("id")

    return None  # Если пользователя не нашли

>>>>>>> 116c7b8 (Первый коммит)
@app.route("/send_message", methods=["POST"])
def send_message_route():
    try:
        # Получаем JSON-данные с именем пользователя и сообщением
        data = request.get_json()
<<<<<<< HEAD
=======
        print(data)
>>>>>>> 116c7b8 (Первый коммит)

        username = data.get("username")  # Имя пользователя
        print(username)
        message = data.get("message")    # Сообщение

        if not username or not message:
            return jsonify({"error": "Отсутствуют данные: username или message"}), 400

        # Получаем ID пользователя
<<<<<<< HEAD
        user_id = 1840233118#ПОЛУЧЕНИЕ АЙДИ ЧЕРЕЗ ИМЯ= app.get_users(username) НЕ РАБОТАЕТ
=======
        #user_id = 1840233118#ПОЛУЧЕНИЕ АЙДИ ЧЕРЕЗ ИМЯ= app.get_users(username) НЕ РАБОТАЕТ
        user_id = get_user_id_by_username(username)
>>>>>>> 116c7b8 (Первый коммит)
        if not user_id:
            return jsonify({"error": f"Не найден пользователь с именем {username}"}), 404

        # Отправляем сообщение через Telegram-бота
        result = send_message(user_id, message)
        return jsonify(result)

    except Exception as e:
        return jsonify({"error": f"Ошибка сервера: {str(e)}"}), 500






def run_flask():
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)




















with open(DATA_FILE, "r") as file:
    config = json.load(file)

def get_current_time_kiev():
    kiev_tz = pytz.timezone('Europe/Kiev')
    now = datetime.now(kiev_tz)
    return now.strftime("%H:%M; %d/%m/%Y")

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_sent_messages():
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)
    return data.get("sent_messages", {})

def save_sent_messages(sent_messages):
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)
    data["sent_messages"] = sent_messages
    with open(DATA_FILE, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

def load_muted_users_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    muted_users = {}
    for user in data.get("users", []):
        if user.get("mute", False):
            mute_end = user.get("mute_end")
            if mute_end:
                mute_end = datetime.strptime(mute_end, "%H:%M; %d/%m/%Y")
            muted_users[user["id"]] = {
                "first_name": user.get("first_name"),
                "second_name": user.get("second_name"),
                "username": user.get("username"),
                "expiration": mute_end,
                "reason": user.get("reason")
            }
    return muted_users

def load_users_info(json_file=DATA_FILE):
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get("users", [])
    except FileNotFoundError:
        print(f"Помилка: Файл '{json_file}' не знайден.")
        return []
    except json.JSONDecodeError:
        print("Помилка: некорректний формат JSON.")
        return []

def load_chat_id_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    chat_id = data.get("chat_id")
    return chat_id

def load_bottocen_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    bot_token = data.get("bot_token")
    return bot_token

def update_data_json(data):
    with open(DATA_FILE, "w") as file:
        json.dump(data, file, indent=4, ensure_ascii=False)


users_info = load_users_info()
muted_users = load_muted_users_from_file()

CREATOR_CHAT_ID = load_chat_id_from_file()
BOTTOCEN = load_bottocen_from_file()



async def start(update: Update, context):
    user = update.message.from_user
    chat_id = update.effective_chat.id

    if chat_id == CREATOR_CHAT_ID:
        await update.message.reply_text("Команда /start недоступна в цій групі.")
        return

    user_found = False
    for u in config["users"]:
        if u["id"] == str(user.id):
            user_found = True
            break

    if not user_found:
        new_user = {
            "id": str(user.id),
            "username": user.username or "Не вказано",
            "first_name": user.first_name or "Не вказано",
            "second_name": user.first_name or "Не вказано",
            "join_date": get_current_time_kiev(),
            "rating": 0,
            "mute": False,
            "mute_end": None,
            "reason": None
        }
        config["users"].append(new_user)
        save_data(config)

    # Загружаем данные чатов
    chats_data = load_chats()

    # Если пользователя нет в чате, добавляем его
    if str(user.id) not in chats_data:
        chats_data[str(user.id)] = {
            "username": user.username or "Не вказано",
            "messages": []  # Список сообщений для этого пользователя
        }
        save_chats(chats_data)

    keyboard = [
        ["/start", "/rate"],
        ["/message", "/stopmessage"],
        ["/fromus", "/help"],
    ]

    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        "Привіт! Я ваш бот підтримки. Введіть команду /rate для оцінки бота, /message для написання адміністраторам бота або /help для отримання інформації про команди.",
        reply_markup=reply_markup
    )

async def rate(update: Update, context):
    user_id = update.message.from_user.id

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    user_rating = None
    for user in data.get("users", []):
        if user.get('id') == str(user_id):
            user_rating = user['rating']
            break

    total_score = data.get("total_score", 0)
    num_of_ratings = data.get("num_of_ratings", 0)

    average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

    rating_text = f"Загальна оцінка: {round(average_rating, 1)}⭐️\nВаш попередній відгук: {user_rating}⭐️"

    keyboard = [
        [InlineKeyboardButton("0.5⭐️", callback_data='0.5'), InlineKeyboardButton("1⭐️", callback_data='1')],
        [InlineKeyboardButton("1.5⭐️", callback_data='1.5'), InlineKeyboardButton("2⭐️", callback_data='2')],
        [InlineKeyboardButton("2.5⭐️", callback_data='2.5'), InlineKeyboardButton("3⭐️", callback_data='3')],
        [InlineKeyboardButton("3.5⭐️", callback_data='3.5'), InlineKeyboardButton("4⭐️", callback_data='4')],
        [InlineKeyboardButton("4.5⭐️", callback_data='4.5'), InlineKeyboardButton("5⭐️", callback_data='5')],
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(f"{rating_text}\nОберіть оцінку:", reply_markup=reply_markup)

async def button_callback(update: Update, context):
    query = update.callback_query
    user_id = query.from_user.id
    new_rating = float(query.data)

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    user_found = False
    previous_rating = 0

    for user in data.get("users", []):
        if user.get('id') == str(user_id):
            previous_rating = user.get('rating', 0)
            user['rating'] = new_rating
            user_found = True
            break

    if not user_found:
        new_user = {
            'id': str(user_id),
            'first_name': query.from_user.first_name,
            'username': query.from_user.username,
            'join_date': datetime.now().strftime("%H:%M; %d/%m/%Y"),
            'rating': new_rating,
            'mute': False,
            'mute_end': None,
            'reason': None
        }
        data['users'].append(new_user)

    total_score = data.get("total_score", 0)
    num_of_ratings = data.get("num_of_ratings", 0)

    if previous_rating == 0:
        num_of_ratings += 1
        total_score += new_rating
    else:
        total_score = total_score - previous_rating + new_rating

    data["total_score"] = total_score
    data["num_of_ratings"] = num_of_ratings

    with open(DATA_FILE, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=5,
                  default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj, datetime) else None)

    average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

    await query.edit_message_text(
        f"Дякуємо за ваш відгук! Ваша оцінка: {new_rating}⭐️\nЗагальна оцінка: {round(average_rating, 1)}⭐️"
    )

async def button(update: Update, context):
    global total_score, num_of_ratings

    query = update.callback_query
    await query.answer()

    selected_rate = float(query.data)

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    total_score = data.get("total_score", 0) + selected_rate
    num_of_ratings = data.get("num_of_ratings", 0) + 1

    data["total_score"] = total_score
    data["num_of_ratings"] = num_of_ratings

    with open(DATA_FILE, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4,
                  default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj, datetime) else None)

    average_rating = total_score / num_of_ratings

    user_id = query.from_user.id
    if user_id in users_info:
        users_info[user_id]['rating'] = selected_rate

    await query.edit_message_text(
        f"Дякуємо за ваш відгук! Ваша оцінка: {selected_rate}⭐️\nЗагальна оцінка: {round(average_rating, 1)}⭐️")

async def auto_delete_message(bot, chat_id, message_id, delay):
    await asyncio.sleep(delay)
    await bot.delete_message(chat_id=chat_id, message_id=message_id)

async def message(update: Update, context):
    user_id = update.message.from_user.id
    muted_users = load_muted_users_from_file()


    if user_id in muted_users and muted_users[user_id]['expiration'] > datetime.now():
        reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
        return

    reply = await update.message.reply_text(
        "Введіть ваше повідомлення, і його буде відправлено адміністраторам бота. Введіть /stopmessage, щоб завершити введення повідомлень."
    )

    context.user_data['waiting_for_message'] = True

    await asyncio.create_task(
        auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))

async def stopmessage(update: Update, context):
    if context.user_data.get('waiting_for_message'):
        reply = await update.message.reply_text("Ви завершили введення повідомлень.")
        context.user_data['waiting_for_message'] = False
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
    else:
        await update.message.reply_text("Ви не в режимі введення повідомлень.")

async def help(update: Update, context):
    if str(update.message.chat.id) == str(CREATOR_CHAT_ID):
        help_text = (
            "Доступні команди в групі:\n"
            "Відповісти на повідомлення бота - Надіслати повідомлення користувачу, який надіслав це повідомлення.\n"
            "/mute <час> <користувач> 'причина' - Замутити користувача на вказаний час.\n"
            "/unmute <користувач> - Розмутити користувача.\n"
            "/mutelist - Показати список замучених користувачів.\n"
            "/alllist - Показати всіх користувачів.\n"
            "/allmessage <повідомлення> - Надіслати повідомлення всім користувачам.\n"
            "/fromus - Інформація про створювача.\n"
            "/help - Показати доступні команди.\n"
            "/info - Показати інформацію про програмістів та адміністраторів.\n"
            "/admin <користувач> - Додати адміністратора.\n"
            "/deleteadmin <користувач> - Видалити адміністратора.\n"
            "/programier <користувач> - Додати програміста.\n"
            "/deleteprogramier <користувач> - Видалити програміста.\n"
            "/get_alllist - Отримати Exel файл з користувачами.\n"
            "/set_alllist - Записати Exel файл з користувачами.\n"
        )
    elif str(update.message.chat.id) == str(-1002358066044):
        help_text = (
            "Доступні команди в групі:\n"
            "/get_alllist - Отримати Exel файл з користувачами.\n"
            "/set_alllist - Записати Exel файл з користувачами.\n"
        )
    else:
        help_text = (
            "Доступні команди в боті:\n"
            "/start - Запустити бота.\n"
            "/rate - Залишити відгук.\n"
            "/message - Почати введення повідомлень адміністраторам.\n"
            "/stopmessage - Завершити введення повідомлень.\n"
            "/fromus - Інформація про створювача.\n"
            "/help - Показати доступні команди.\n"
        )

    await update.message.reply_text(help_text)

async def fromus(update: Update, context):
    await update.message.reply_text(
        "*Skeleton*  Написв бота\nПортфоліо:  ```https://www.linkedin.com/in/artem-k-972a41344/``` \n Телеграм канал з усіма проєктами: ```https://t.me/AboutMyProjects```\n По всім питанням пишіть в цього бота",
        parse_mode="MarkdownV2"
    )

async def info(update: Update, context: CallbackContext):
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    programmers = data.get("programmers", [])
    admins = data.get("admins", [])

    programmer_list = "\n".join(programmers) if programmers else "Список программистов пуст."
    admin_list = "\n".join(admins) if admins else "Список администраторов пуст."

    await update.message.reply_text(f"Программісти:\n{programmer_list}\n\nАдміністратори:\n{admin_list}")


async def update_website(message_info):
    url = "http://127.0.0.1:5000/update_chat"  # Отправляем локально
    headers = {"Content-Type": "application/json"}

    try:
        response = requests.post(url, json=message_info, headers=headers)
        if response.status_code == 200:
            print("✅ Данные успешно отправлены на сайт")
        else:
            print(f"❌ Ошибка {response.status_code}: {response.text}")
    except Exception as e:
        print(f"Ошибка при отправке данных на сайт: {e}")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sent_messages = load_sent_messages()
    muted_users = load_muted_users_from_file()
    if context.user_data.get("awaiting_file"):
        if update.message.document:
            document = update.message.document
            file_path = "uploaded_file.xlsx"

            file = await document.get_file()
            await file.download_to_drive(file_path)

            try:
                wb = load_workbook(file_path)

                sheet_all_user = wb["AllUser"]
                sheet_admins = wb["Admins"]
                sheet_programmers = wb["Programmers"]
                sheet_general_info = wb["GeneralInfo"]
                sheet_sent_messages = wb["SentMessages"]

                updated_users = []
                muted_users = {}
                sent_messages = {}

                for row in sheet_all_user.iter_rows(min_row=2, values_only=True):
                    if len(row) < 9:
                        continue

                    user_data = {
                        "id": str(row[0]),
                        "first_name": row[1],
                        "second_name": row[2],
                        "username": row[3],
                        "join_date": row[4].strftime("%H:%M; %d/%m/%Y") if isinstance(row[3], datetime) else str(row[4]),
                        "rating": int(row[5]) if row[5] is not None else 0,
                        "mute": bool(row[6]),
                        "mute_end": row[7].strftime("%H:%M; %d/%m/%Y") if isinstance(row[6], datetime) else str(row[7]),
                        "reason": row[8]
                    }
                    updated_users.append(user_data)

                    if user_data["mute"]:
                        muted_users[user_data["username"]] = True

                for row in sheet_sent_messages.iter_rows(min_row=2, values_only=True):
                    if len(row) < 2 or not row[0] or not row[1]:
                        continue
                    sent_messages[str(row[0])] = row[1]

                admins = [row[0] for row in sheet_admins.iter_rows(min_row=2, values_only=True)]
                programmers = [row[0] for row in sheet_programmers.iter_rows(min_row=2, values_only=True)]

                bot_token = sheet_general_info.cell(row=2, column=1).value or ""
                owner_id = sheet_general_info.cell(row=2, column=2).value or ""
                chat_id = sheet_general_info.cell(row=2, column=3).value or ""
                total_score = float(sheet_general_info.cell(row=2, column=4).value or 0)
                num_of_ratings = int(sheet_general_info.cell(row=2, column=5).value or 0)

                data = {
                    "users": updated_users,
                    "muted_users": muted_users,
                    "admins": admins,
                    "programmers": programmers,
                    "bot_token": bot_token,
                    "owner_id": owner_id,
                    "chat_id": chat_id,
                    "total_score": total_score,
                    "num_of_ratings": num_of_ratings,
                    "sent_messages": sent_messages,
                }

                with open(DATA_FILE, "w", encoding="utf-8") as json_file_obj:
                    json.dump(data, json_file_obj, ensure_ascii=False, indent=4,
                              default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj,
                                                                                                datetime) else None)

                await update.message.reply_text("Файл успешно обработан!")

            except Exception as e:
                await update.message.reply_text(f"Помилка при обробці файла: {e}")

            finally:
                context.user_data["awaiting_file"] = False
        else:
            await update.message.reply_text("Пожалуйста, отправьте Excel-файл.")
    elif (str(update.message.chat.id)) != (str(CREATOR_CHAT_ID)):
        user_id = update.message.from_user.id
        if user_id in muted_users and muted_users[user_id]['expiration'] > datetime.now():
            reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

        if context.user_data.get('waiting_for_message'):
            user_name = update.effective_user.first_name
            user_username = update.effective_user.username if update.effective_user.username else "немає імені користувача"
            current_time = get_current_time_kiev()
            user_message = update.message.text if update.message.text else ""

            first_message = f'Повідомлення від **{user_name}**; ```@{user_username}``` \n{current_time}:'
            if user_message:
                first_message += f'\n{user_message}'

            # Загружаем данные чатов из chats.json
            chats_data = load_chats()
            chat_id_str = str(update.message.chat.id)

            # Проверяем, существует ли чат, если нет — создаем
            if chat_id_str not in chats_data or not isinstance(chats_data[chat_id_str], dict):
                print(f"Ошибка: данные для чата {chat_id_str} повреждены, исправляем.")
                chats_data[chat_id_str] = {"username": user_username, "messages": []}

            # Проверяем, что "messages" существует и это список
            if "messages" not in chats_data[chat_id_str] or not isinstance(chats_data[chat_id_str]["messages"], list):
                print(f"Ошибка: messages в чате {chat_id_str} повреждены, исправляем.")
                chats_data[chat_id_str]["messages"] = []

            # Сохраняем информацию о сообщении (без user_id)
            message_info = {
                "username": user_username,
                "message": user_message,
                "time_sent": current_time
            }

            # Добавляем сообщение в список
            chats_data[chat_id_str]["messages"].append(message_info)

            # Сохраняем обновленные данные в chats.json
            save_chats(chats_data)

            # Обновление информации на сайте
            await update_website(message_info)  # Тут будет исправленный запрос (см. ниже)

            # Ответ пользователю
            reply = await update.message.reply_text("Ваше повідомлення надіслано адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
    else:
        if update.effective_user.id != context.bot.id:
            if update.message.reply_to_message:
                if update.message.reply_to_message.from_user.id == context.bot.id:
                    original_message_id = str(update.message.reply_to_message.message_id)
                    if original_message_id in sent_messages:
                        original_user_id = sent_messages[original_message_id]
                        reply_text = update.message.text if update.message.text else ""
                        for user in config['users']:
                            if str(user['id']) == str(original_user_id):
                                user_name = user['first_name']
                                break

                        if update.message.photo:
                            photo_file_id = update.message.photo[-1].file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_photo(chat_id=original_user_id, photo=photo_file_id, caption=caption)

                        elif update.message.document:
                            document_file_id = update.message.document.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_document(chat_id=original_user_id, document=document_file_id, caption=caption)
                        elif update.message.sticker:
                            sticker_file_id = update.message.sticker.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_sticker(chat_id=original_user_id, sticker=sticker_file_id)

                        elif update.message.voice:
                            voice_file_id = update.message.voice.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_voice(chat_id=original_user_id, voice=voice_file_id, caption=caption)

                        elif update.message.video:
                            video_file_id = update.message.video.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_video(chat_id=original_user_id, video=video_file_id, caption=caption)

                        elif update.message.video_note:
                            video_note_file_id = update.message.video_note.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_video_note(chat_id=original_user_id, video_note=video_note_file_id)
                        else:
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_message(chat_id=original_user_id, text=reply_text)
                        await update.message.reply_text(f"Користувачу { user_name } було надіслано повідомлення")
                        sent_messages[update.message.message_id] = update.message.from_user.id
                        save_sent_messages(sent_messages)


async def mute(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    message_text = update.message.text.split()

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return

    mute_time = 300
    reason = "По рішенню адміністратора"
    username = None

    if len(context.args) > 0:
        if context.args[0].isdigit():
            mute_time = int(context.args[0])
            username = context.args[1].lstrip('@') if len(context.args) > 1 else None
        else:
            username = context.args[0].lstrip('@')

    reason_match = re.search(r'["\'](.*?)["\']', update.message.text)
    if reason_match:
        reason = reason_match.group(1)

    if not username:
        await update.message.reply_text("Не вказано користувача для мута.")
        return

    user = next((u for u in config["users"] if u["username"].lower() == username.lower() or str(u["id"]) == username),
                None)

    if not user:
        await update.message.reply_text(f"Користувач {username} не знайден.")
        return

    if user["id"] == config["owner_id"]:
        await update.message.reply_text("Неможливо замутити власника чату.")
        return

    if user["mute"]:
        await update.message.reply_text(f"Користувач {user['first_name']} вже був замучений.")


    user["mute"] = True
    user["mute_end"] = (datetime.now() + timedelta(seconds=mute_time)).strftime("%H:%M; %d/%m/%Y")
    user["reason"] = reason

    config["muted_users"][username] = True
    save_data(config)

    mute_permissions = ChatPermissions(can_send_messages=False)
    await context.bot.restrict_chat_member(chat_id=config["chat_id"], user_id=user["id"], permissions=mute_permissions)
    await context.bot.send_message(chat_id=user["id"],
                                   text=f"Вас замутили на {str(timedelta(seconds=mute_time))}\nПричина: {reason}")
    await update.message.reply_text(f"Користувач @{user['username']} замучений.")

async def unmute(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("Використовуйте: /unmute <користувач>")
        return



    username = context.args[0].lstrip('@')

    user = next((u for u in config["users"] if u["username"].lower() == username.lower() or str(u["id"]) == username), None)

    if user and user["mute"]:
        user["mute"] = False
        user["mute_end"] = None
        user["reason"] = None

        config["muted_users"].pop(username, None)
        save_data(config)

        mute_permissions = ChatPermissions(can_send_messages=True)
        await context.bot.restrict_chat_member(chat_id=config["chat_id"], user_id=user["id"], permissions=mute_permissions)
        await update.message.reply_text(f"Користувач @{user['username']} був розмучений.")
    else:
        await update.message.reply_text(f"Користувач {username} не знайден або не був замучений.")

async def admin(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user):
        await update.message.reply_text("Ця команда доступна тільки програмістам.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("Використовуйте: /admin @username")
        return

    username = context.args[0].lstrip('@')
    if username in config["admins"]:
        await update.message.reply_text(f"Користувач @{username} вже є администратором.")
    else:
        config["admins"].append(username)
        save_data(config)
        await update.message.reply_text(f"Користувач @{username} додан в список администраторів.")

async def deleteadmin(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user):
        await update.message.reply_text("Ця команда доступна тільки програмістам.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("Використовуйте: /deleteadmin @username")
        return

    username = context.args[0].lstrip('@')
    if username in config["admins"]:
        config["admins"].remove(username)
        save_data(config)
        await update.message.reply_text(f"Користувач @{username} видален зі списку администраторів.")
    else:
        await update.message.reply_text(f"Користувач @{username} не знайден.")

async def programier(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if is_programmer(user):
        if len(context.args) > 0:
            new_programmer = context.args[0].replace("@", "")
            if new_programmer not in config["programmers"]:
                config["programmers"].append(new_programmer)
                save_data(config)
                await update.message.reply_text(f"Користувач {new_programmer} додан в список программістів.")
            else:
                await update.message.reply_text(f"Користувач {new_programmer} вже є в списку программистів.")
        else:
            await update.message.reply_text("Використовуйте: /programier @username")
    else:
        await update.message.reply_text("Ця команда доступна лише адміністраторам.")

async def deleteprogramier(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if is_programmer(user):
        if len(context.args) > 0:
            removed_programmer = context.args[0].replace("@", "")
            if removed_programmer == "ArtemKirss":
                await update.message.reply_text(f"Неможливо видалити {removed_programmer} зі списку программистов.")
            elif removed_programmer in config["programmers"]:
                config["programmers"].remove(removed_programmer)
                save_data(config)
                await update.message.reply_text(f"Користувач {removed_programmer} видален зі списку программистів.")
            else:
                await update.message.reply_text(f"Користувач {removed_programmer} не є программистом.")
        else:
            await update.message.reply_text("Використовуйте: /deleteprogramier @username")
    else:
        await update.message.reply_text("Ця команда доступна лише адміністраторам.")

async def mutelist(update: Update, context):
    user = update.message.from_user.username
    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("Ця команда доступна тільки адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    admins = data.get("admins", [])
    programmers = data.get("programmers", [])
    muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

    response = "Замучені користувачі:\n"

    if muted_users:
        for user_id, mute_info in muted_users.items():
            expiration = mute_info.get('mute_end', 'Невідомо')
            reason = mute_info.get('reason', 'Без причини')

            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=int(user_id))
            user_fullname = user_info.user.first_name or "Невідомий"
            username = user_info.user.username or "Немає імені користувача"

            join_date = mute_info.get('join_date', 'Невідома')
            rating = mute_info.get('rating', 0)
            mute_symbol = "🔇"

            admins_sumdol = "👨🏻‍💼"
            if username in admins:
                admins_sumdol = "👮🏻‍♂️"
            if username in programmers:
                admins_sumdol = "👨🏻‍💻"

            response += (
                f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                f"Залишилось: {expiration}\n"
                f"Причина: {reason}\n"
                f"Дата заходу: {join_date}\n"
                f"Оцінка: {rating}⭐️\n"
                "-------------------------------------------------------------------------\n"
            )
    else:
        response += "Немає замучених користувачів.\n"
        response += "-------------------------------------------------------------------------\n"

    await update.message.reply_text(response)

async def alllist(update: Update, context: CallbackContext):
    global mute_symbol
    user = update.message.from_user.username
    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("Ця команда доступна лише адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    admins = data.get("admins", [])
    programmers = data.get("programmers", [])
    users_info = {user['id']: user for user in data.get("users", [])}
    muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

    response = "Користувачі:\n"
    unique_users = {user['id'] for user in data.get("users", [])}

    if unique_users:
        for user_id in unique_users:
            user_data = users_info.get(str(user_id), {})
            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=user_id)
            user_fullname = user_info.user.first_name or "Невідомий"
            username = user_info.user.username or "Немає імені користувача"
            join_date = user_data.get('join_date', 'Невідома')
            rating = user_data.get('rating', 0)

            admins_sumdol = "👨🏻‍💼"
            if username in admins:
                admins_sumdol = "👮🏻‍♂️"
            if username in programmers:
                admins_sumdol = "👨🏻‍💻"

            mute_symbol = "🔇" if str(user_id) in muted_users else "🔊"

            response += f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\nДата заходу: {join_date}\nОцінка: {rating}⭐️\n"
            response += "-------------------------------------------------------------------------\n"
    else:
        response += "Немає користувачів.\n"
        response += "-------------------------------------------------------------------------\n"

    response += "==========================================\n"
    response += "\n"
    response += "==========================================\n"
    response += "Замучені користувачі:\n"

    if muted_users:
        for user_id, mute_info in muted_users.items():
            expiration = mute_info['mute_end'] or "Невідомо"
            reason = mute_info.get('reason', "Без причини")
            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=user_id)
            user_fullname = user_info.user.first_name or "Невідомий"
            username = user_info.user.username or "Немає імені користувача"
            user_data = users_info.get(str(user_id), {})
            join_date = user_data.get('join_date', 'Невідома')
            rating = user_data.get('rating', 0)

            admins_sumdol = "👨🏻‍💼"
            if username in admins:
                admins_sumdol = "👮🏻‍♂️"
            if username in programmers:
                admins_sumdol = "👨🏻‍💻"

            mute_symbol = "🔇"

            response += (
                f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                f"Залишилось: {str(expiration).split('.')[0]}\n"
                f"Причина: {reason}\n"
                f"Дата заходу: {join_date}\n"
                f"Оцінка: {rating}⭐️\n"
                "-------------------------------------------------------------------------\n"
            )
    else:
        response += "Немає замучених користувачів.\n"
        response += "-------------------------------------------------------------------------\n"

    await update.message.reply_text(response)

async def allmessage(update: Update, context):
    user = update.message.from_user.username

    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("Ця команда доступна тільки адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    if not context.args:
        await update.message.reply_text("Будь ласка, укажіть текст повідомлення після команди.")
        return

    message_text = update.message.text.split(' ', 1)[1]

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        config = json.load(file)

    users = config.get("users", [])

    for user_data in users:
        user_id = user_data.get("id")
        if user_id:
            try:
                await context.bot.send_message(chat_id=user_id, text=message_text)
            except Exception as e:
                print(f"Помилка при відправці повідомлення користувачу {user_id}: {e}")

    await update.message.reply_text("Повідомлення відправлено всім користувачам.")

def is_programmer(username):
    return username in config["programmers"]

def is_admin(username):
    return username in config["admins"]


async def get_alllist(update: Update, context: CallbackContext) -> None:
    user = update.message.from_user.username

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)

        all_users_df = pd.DataFrame(data["users"])

        users_df = all_users_df[all_users_df["mute"] == False]
        muted_df = all_users_df[all_users_df["mute"] == True]

        muted_df.loc[:, "mute_end"] = muted_df["mute_end"].apply(
            lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y") if isinstance(x, str) else ""
        )

        admins_df = pd.DataFrame(data.get("admins", []), columns=["Admins"])
        programmers_df = pd.DataFrame(data.get("programmers", []), columns=["Programmers"])
        general_info_df = pd.DataFrame(
            [{
                "bot_token": data.get("bot_token"),
                "owner_id": data.get("owner_id"),
                "chat_id": data.get("chat_id"),
                "total_score": data.get("total_score"),
                "num_of_ratings": data.get("num_of_ratings")
            }]
        )
        sent_messages_df = pd.DataFrame(data.get("sent_messages", {}).items(), columns=["MessageID", "UserID"])
        muted_users_df = pd.DataFrame(data.get("muted_users", {}).items(), columns=["UserID", "Details"])

        excel_file = "Supp0rts2Bot_all_users.xlsx"
        with pd.ExcelWriter(excel_file) as writer:
            users_df.to_excel(writer, index=False, sheet_name="Users")
            muted_df.to_excel(writer, index=False, sheet_name="Muted")
            all_users_df.to_excel(writer, index=False, sheet_name="AllUser")
            admins_df.to_excel(writer, index=False, sheet_name="Admins")
            programmers_df.to_excel(writer, index=False, sheet_name="Programmers")
            general_info_df.to_excel(writer, index=False, sheet_name="GeneralInfo")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")
            muted_users_df.to_excel(writer, index=False, sheet_name="MutedUsers")

        workbook = load_workbook(excel_file)
        sheet = workbook["AllUser"]

        yellow_fill = PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
        red_fill = PatternFill(start_color="b40a0a", end_color="b40a0a", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
            username_cell = row[3]
            mute_status = next((user['mute'] for user in data["users"] if user["username"] == username_cell.value), False)

            fill_color = red_fill if mute_status else yellow_fill

            for cell in row[:9]:
                cell.fill = fill_color

        workbook.save(excel_file)

        await update.message.reply_document(document=open(excel_file, "rb"))

    except Exception as e:
        await update.message.reply_text(f"Error: {e}")

async def set_alllist(update: Update, context: CallbackContext) -> None:
    user = update.message.from_user.username

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return
    await update.message.reply_text("Будь ласка пришліть Excel file з данними.")
    context.user_data["awaiting_file"] = True

async def set_default_commands(application):
    commands = [
        BotCommand("start", "Запустити бота"),
        BotCommand("rate", "Залишити відгук"),
        BotCommand("message", "Почати введення повідомлень адміністраторам"),
        BotCommand("stopmessage", "Завершити введення повідомлень"),
        BotCommand("fromus", "Інформація про створювача"),
        BotCommand("help", "Показати доступні команди"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeDefault())

async def set_creator_commands(application):
    commands = [
        BotCommand("mutelist", "Показати список замучених користувачів"),
        BotCommand("alllist", "Показати всіх користувачів"),
        BotCommand("fromus", "Інформація про створювача"),
        BotCommand("help", "Показати доступні команди"),
        BotCommand("info", "Показати інформацію про програмістів та адміністраторів"),
        BotCommand("get_alllist", "Отримати Exel файл з користувачами"),
        BotCommand("set_alllist", "Записати Exel файл з користувачами"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=CREATOR_CHAT_ID))

async def set_save_commands(application):
    commands = [
        BotCommand("get_alllist", "Отримати Exel файл з користувачами"),
        BotCommand("set_alllist", "Записати Exel файл з користувачами"),
        BotCommand("help", "Показати доступні команди"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=-1002310142084))

async def send_user_list():
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)

        all_users_df = pd.DataFrame(data["users"])
        print(all_users_df)
        users_df = all_users_df[all_users_df["mute"] == False]
        print(users_df)
        muted_df = all_users_df[all_users_df["mute"] == True]
        print(muted_df)

        muted_df.loc[:, "mute_end"] = muted_df["mute_end"].apply(
            lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y") if isinstance(x, str) else ""
        )

        admins_df = pd.DataFrame(data.get("admins", []), columns=["Admins"])
        programmers_df = pd.DataFrame(data.get("programmers", []), columns=["Programmers"])
        general_info_df = pd.DataFrame(
            [{
                "bot_token": data.get("bot_token"),
                "owner_id": data.get("owner_id"),
                "chat_id": data.get("chat_id"),
                "total_score": data.get("total_score"),
                "num_of_ratings": data.get("num_of_ratings")
            }]
        )
        sent_messages_df = pd.DataFrame(data.get("sent_messages", {}).items(), columns=["MessageID", "UserID"])
        muted_users_df = pd.DataFrame(data.get("muted_users", {}).items(), columns=["UserID", "Details"])

        excel_file = "Supp0rts2Bot_all_users.xlsx"
        with pd.ExcelWriter(excel_file) as writer:
            users_df.to_excel(writer, index=False, sheet_name="Users")
            muted_df.to_excel(writer, index=False, sheet_name="Muted")
            all_users_df.to_excel(writer, index=False, sheet_name="AllUser")
            admins_df.to_excel(writer, index=False, sheet_name="Admins")
            programmers_df.to_excel(writer, index=False, sheet_name="Programmers")
            general_info_df.to_excel(writer, index=False, sheet_name="GeneralInfo")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")
            muted_users_df.to_excel(writer, index=False, sheet_name="MutedUsers")

        workbook = load_workbook(excel_file)
        sheet = workbook["AllUser"]

        yellow_fill = PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
        red_fill = PatternFill(start_color="b40a0a", end_color="b40a0a", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8):
            username_cell = row[2]
            mute_status = next((user['mute'] for user in data["users"] if user["username"] == username_cell.value), False)

            fill_color = red_fill if mute_status else yellow_fill

            for cell in row[:8]:
                cell.fill = fill_color

        workbook.save(excel_file)

        bot = Bot(token=BOTTOCEN)
        await bot.send_document(chat_id=-1002358066044, document=open(excel_file, "rb"))

    except Exception as e:
        bot = Bot(token=BOTTOCEN)
        await bot.send_message(chat_id=-1002358066044, text=f"Ошибка при создании отчета: {e}")

async def main():
    application = Application.builder().token("7677888606:AAHMm3aSt84ZQkJ0wrlH4__St3lW36-TL8g").build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("rate", rate))
    application.add_handler(CommandHandler("message", message))
    application.add_handler(CommandHandler("stopmessage", stopmessage))
    application.add_handler(CommandHandler("fromus", fromus))
    application.add_handler(CommandHandler("help", help))
    application.add_handler(CommandHandler("mute", mute))
    application.add_handler(CommandHandler("unmute", unmute))
    application.add_handler(CommandHandler("mutelist", mutelist))
    application.add_handler(CommandHandler("alllist", alllist))
    application.add_handler(CommandHandler("allmessage", allmessage))
    application.add_handler(CommandHandler("admin", admin))
    application.add_handler(CommandHandler("deleteadmin", deleteadmin))
    application.add_handler(CommandHandler("programier", programier))
    application.add_handler(CommandHandler("deleteprogramier", deleteprogramier))
    application.add_handler(CommandHandler("info", info))
    application.add_handler(CommandHandler("get_alllist", get_alllist))
    application.add_handler(CommandHandler("set_alllist", set_alllist))

    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(CallbackQueryHandler(button))
    application.add_handler(MessageHandler(filters.ALL, handle_message))

    await set_default_commands(application)
    await set_creator_commands(application)
    await set_save_commands(application)


    scheduler = AsyncIOScheduler(timezone=pytz.timezone("Europe/Kyiv"))
    scheduler.add_job(send_user_list, "cron", hour=0, minute=0)
    scheduler.start()

    application.run_polling()


if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()
    asyncio.run(main())

